
import pandas as pd
import pysam
import numpy as np
import re
import openpyxl
from openpyxl.styles import numbers,PatternFill
import string
from pathlib import Path
from scipy.sparse import coo_matrix
from collections import OrderedDict

import myRiboSeq.mylib_bin as my
import myRiboSeq.myRiboBin as mybin

# thresholds for total counts of transcripts
thresholds = [np.inf,64,32,16,8,0]

dict_pattern = {
    'hsa':'ENST([0-9]{11})',
    'mmu':'ENSMUST([0-9]{11})',
    'sc':'YAL([0-9A-Z-])'
}
dict_format = {
    'hsa':'ENST{:0>11}',
    'mmu':'ENSMUST{:0>11}',
    'sc':'YAL{%s}'
}

dict_nt_bin = {'A': 0, 'G': 1, 'C': 2, 'T': 3, 'N': 4}
nt_list = ['A', 'G', 'C', 'T', 'N']

biotype_list = ['protein-coding','ncrna','genome','pseudogene']

'''preparation of data for the downstream analyses'''
def prep_data(
    save_dir,
    ref_dir,
    data_dir,
    sp):

    ref = my.Ref(
        data_dir=data_dir,
        ref_dir=ref_dir,
        sp=sp)

    if save_dir is None:
        return ref
        
    save_dir = save_dir / 'prep_data'
    if not save_dir.exists():
        save_dir.mkdir()

    n_reads_list = [];n_reads_full_list = []
    for a in ref.exp_metadata.df_metadata['align_files']:
        smpl_name = ref.exp_metadata.df_metadata.query(f'align_files == "{a}"').sample_name.iloc[-1]
        print(f'preparing data for {smpl_name}...')
        infile = pysam.AlignmentFile(ref.data_dir / a)

        obj = mybin.myBinRibo(
            data_dir=data_dir,
            smpl=smpl_name,
            sp=sp,
            save_dir=save_dir
        )
        obj.encode(infile,ref)
        n_reads_list.append(obj.n_reads)
        obj.decode()
        df,_ = obj.make_df(tr_cnt_thres=-1,is_frame=False,is_cds=False,is_seq=False)
        idx_full = df['length'] == df['read_length']
        n_reads_full_list.append(np.sum(idx_full))
        print(f"total reads: {obj.n_reads}...")
        print(f"fully aligned reads: {np.sum(idx_full)}...")

    pd.DataFrame({
        'total_reads':n_reads_list,
        'fully_aligned_reads':n_reads_full_list
    },index=ref.exp_metadata.df_metadata['sample_name']).\
        to_csv(save_dir / 'summary.csv.gz')

    return ref

def prep_data_genome(
    save_dir:Path,
    ref_dir:Path,
    ref_dir_genome:Path,
    data_dir:Path,
    sp:str,
    biotypes=[],
    sense_antisense=False):

    assert np.all([b in biotype_list for b in biotypes ])

    refg = my.RefGenome(
        data_dir=data_dir,
        ref_dir=ref_dir,
        ref_dir_genome=ref_dir_genome,
        sp=sp,
        biotypes=['protein-coding','ncrna','pseudogene'])

    if save_dir is None:
        return refg

    
    save_dir = save_dir / 'prep_data'
    if not save_dir.exists():
        save_dir.mkdir()

    dict_n_reads_list = {}
    for a in refg.exp_metadata.df_metadata['align_files']:
        smpl_name = refg.exp_metadata.df_metadata.query(f'align_files == "{a}"').sample_name.iloc[-1]
        print(f'preparing data for {smpl_name}...')
        infile = pysam.AlignmentFile(refg.data_dir / a)

        if sense_antisense:
            for strand in ['sense','antisense']:
                obj = mybin.myBinRiboGenome(
                    data_dir=data_dir,
                    smpl=smpl_name,
                    sp=sp,
                    save_dir=save_dir
                )
                obj.encode(infile,refg,is_sense=True if strand == 'sense' else False)
                # obj.decode(mode='tr',data_path=save_dir / f'dat_{smpl_name}_ncrna.bin.gz')
                # df,_ = obj.make_df(tr_cnt_thres=8,is_frame=True)
                dict_n_reads_list[smpl_name + '_' + strand] = obj.dict_n_reads
                dict_n_reads_list[smpl_name + '_' + strand]['intergenic'] = obj.n_reads - np.sum(list(obj.dict_n_reads.values()))

        else:
            obj = mybin.myBinRiboGenome(
                data_dir=data_dir,
                smpl=smpl_name,
                sp=sp,
                save_dir=save_dir
            )
            obj.encode(infile,refg)
            # obj.decode(mode='tr',data_path=save_dir / f'dat_{smpl_name}_ncrna.bin.gz')
            # df,_ = obj.make_df(tr_cnt_thres=8,is_frame=True)
            dict_n_reads_list[smpl_name] = obj.dict_n_reads
            dict_n_reads_list[smpl_name]['intergenic'] = obj.n_reads - np.sum(list(obj.dict_n_reads.values()))

    # pd.DataFrame(n_reads_list,index=ref.exp_metadata.df_metadata['sample_name'],columns=['total_reads']).\
    #     to_csv(save_dir / 'summary.csv.gz')
    # output read number information
    smpl_list = list(dict_n_reads_list.keys())
    smpls = '\t'.join(smpl_list)

    with open(save_dir / 'summary_reads.csv', 'w' ) as f:
        f.write(f'Mapping to genome\t{smpls}\n')
        s = smpl_list[0]
        attrs_list = list(dict_n_reads_list[s].keys())
        attrs = [
            '_'.join(x) 
            if x != 'intergenic' else x
            for x in dict_n_reads_list[s].keys()
            ]
        for a_,a in zip(attrs,attrs_list):
            vals = '\t'.join([str(dict_n_reads_list[s][a]) for s in smpl_list])
            f.write(f'{a_}\t{vals}\n')


def prep_data_3step(
    save_dir,
    ref_dir,
    ref_dir_genome,
    data_dir,
    sp,
    biotypes:list):

    assert np.all([b in biotype_list for b in biotypes ])

    out = []

    if 'protein-coding' in biotypes:
        ref = my.Ref(
            data_dir=data_dir,
            ref_dir=ref_dir,
            sp=sp)
        if save_dir is None:
            out.append(ref)
    
    if 'ncrna' in biotypes:
        refnc = my.RefNC(
            data_dir=data_dir,
            ref_dir=ref_dir_genome,
            sp=sp)
        if save_dir is None:
            out.append(refnc)
    
    if 'pseudogene' in biotypes:
        refp = my.RefPseudo(
            data_dir=data_dir,
            ref_dir=ref_dir_genome,
            sp=sp)
        if save_dir is None:
            out.append(refp)
    
    if 'genome' in biotypes:
        refg = my.RefGenome(
            data_dir=data_dir,
            ref_dir=ref_dir,
            ref_dir_genome=ref_dir_genome,
            sp=sp,
            biotypes=biotypes)
        if save_dir is None:
            out.append(refg)
    
    if save_dir is None:
        return out
    
    save_dir = save_dir / 'prep_data'
    if not save_dir.exists():
        save_dir.mkdir()

    dict_n_reads_list = {}
    n_pc_reads_list = {}; n_ncrna_reads_list = {}
    for s in ref.exp_metadata.df_metadata['sample_name']:

        # mapping results for protein-coding genes
        a = ref.exp_metadata.df_metadata.query(f'sample_name == "{s}"').align_files.iloc[-1]
        print(f'preparing data for {s}...')
        infile = pysam.AlignmentFile(ref.data_dir / a)

        obj = mybin.myBinRibo(
            data_dir=data_dir,
            smpl=s,
            sp=sp,
            save_dir=save_dir
        )
        obj.encode(infile,ref)
        # obj.decode(mode='tr',data_path=save_dir / f'dat_{smpl_name}_ncrna.bin.gz')
        # df,_ = obj.make_df(tr_cnt_thres=8,is_frame=True)
        n_pc_reads_list[s] = obj.n_reads

        # mapping results for ncRNA genes
        a = ref.exp_metadata.df_metadata.query(f'sample_name == "{s}"').align_ncrna_files.iloc[-1]
        print(f'preparing data for {s} mapped to ncRNAs...')
        infile = pysam.AlignmentFile(ref.data_dir / a)

        obj = mybin.myBinRiboNC(
            data_dir=data_dir,
            smpl=s,
            sp=sp,
            save_dir=save_dir,
            biotype='ncrna'
        )
        obj.encode(infile,refnc)
        # obj.decode(mode='tr',data_path=save_dir / f'dat_{smpl_name}_ncrna.bin.gz')
        # df,_ = obj.make_df(tr_cnt_thres=8,is_frame=True)
        n_ncrna_reads_list[s] = obj.n_reads


        # mapping results for genomes
        a = ref.exp_metadata.df_metadata.query(f'sample_name == "{s}"').align_genome_files.iloc[-1]
        print(f'preparing data for {s} mapped to genome...')
        infile = pysam.AlignmentFile(ref.data_dir / a)

        obj = mybin.myBinRiboGenome(
            data_dir=data_dir,
            smpl=s + '_',
            sp=sp,
            save_dir=save_dir
        )
        obj.encode(infile,refg)
        dict_n_reads_list[s] = obj.dict_n_reads
        dict_n_reads_list[s]['intergenic'] = obj.n_reads - np.sum(list(obj.dict_n_reads.values()))

    # output read number information
    smpl_list = list(n_pc_reads_list.keys())
    smpls = '\t'.join(smpl_list)
    with open(save_dir / 'summary_reads.csv', 'w' ) as f:
        f.write(f'Mapping to transcriptome (protein-coding genes)\t{smpls}\n')
        vals = '\t'.join([str(x) for x in n_pc_reads_list.values()])
        f.write(f'\t{vals}\n')

        f.write(f'Mapping to transcriptome (ncRNAs)\t{smpls}\n')
        vals = '\t'.join([str(x) for x in n_ncrna_reads_list.values()])
        f.write(f'\t{vals}\n')
        
        f.write(f'Mapping to genome\t{smpls}\n')
        attrs_list = list(dict_n_reads_list[s].keys())
        attrs = ['_'.join(x) for x in dict_n_reads_list[smpl_name].keys()]
        for a_,a in zip(attrs,attrs_list):
            vals = '\t'.join([str(dict_n_reads_list[s][a]) for s in smpl_list])
            f.write(f'{a_}\t{vals}\n')

    print("hoge")
    # pd.DataFrame(n_reads_list,index=ref.exp_metadata.df_metadata['sample_name'],columns=['total_reads']).\
    #     to_csv(save_dir / 'summary.csv.gz')

    return ref

    
def _calc_norm_ribo_density(
    df_data,
    dict_tr,
    read_len,
    mode,
    ref,
    is_norm):
    # ribosomde density from start codon
    # read count: [position] * [transcript]
    dict_encode = {
        'count5':[],
        'tr5':[],
        'pos5':[],
        'count3':[],
        'tr3':[],
        'pos3':[]
        }
    for end in [5,3]:
        if mode in ('start','stop'):
            idx = ((df_data[f'dist{end}_start'] > -600) *\
                (df_data[f'dist{end}_stop'] < 600) *\
                (df_data['length'] >= read_len[0]) *\
                (df_data['length'] <= read_len[1])).values
            # idx = ((df_data[f'dist{end}_start'] > -600) *\
            #     (df_data[f'dist{end}_stop'] < 600) *\
            #     (df_data['read_length'] >= read_len[0]) *\
            #     (df_data['read_length'] <= read_len[1])).values
            count5_pos = df_data.iloc[idx,:][["tr_id",f"dist{end}_{mode}"]]\
                .pivot_table(index=f"dist{end}_{mode}",columns="tr_id",aggfunc=len,fill_value=0)
        
        elif mode == '3end':
            tr_len_list = np.array([
                ref.annot.annot_dict[tr]["cdna_len"]
                for tr in df_data["tr_id"]
            ])
            df_data[f'dist{end}_start2'] = df_data[f'dist{end}_start'] + df_data['start'] - tr_len_list
            count5_pos = df_data.query((
                f'(length >= {read_len[0]}) and '
                f'(length <= {read_len[1]})'))[["tr_id",f"dist{end}_start2"]]\
                .pivot_table(index=f"dist{end}_start2",columns="tr_id",aggfunc=len,fill_value=0)
        
        elif mode == '5end':
            tr_start_list = np.array([
                ref.annot.annot_dict[tr]["start"]
                for tr in df_data["tr_id"]
            ])
            df_data[f'dist{end}_start2'] = df_data[f'dist{end}_start'] + tr_start_list
            count5_pos = df_data.query((
                f'(length >= {read_len[0]}) and '
                f'(length <= {read_len[1]})'))[["tr_id",f"dist{end}_start2"]]\
                .pivot_table(index=f"dist{end}_start2",columns="tr_id",aggfunc=len,fill_value=0)

            
        if len(count5_pos) > 0:
            if is_norm:
                # normalize for each gene -> normalize for each position
                tr_norm_term = np.array([
                    np.sum(
                        (dict_tr[tr]['cds_label'] == 1)*\
                        (dict_tr[tr]['length'] >= read_len[0] )*\
                        (dict_tr[tr]['length'] <= read_len[1] )
                    ) / ref.annot.annot_dict[tr]["cds_len"]
                    if np.sum(
                        (dict_tr[tr]['cds_label'] == 1)*\
                        (dict_tr[tr]['length'] >= read_len[0] )*\
                        (dict_tr[tr]['length'] <= read_len[1] )
                    )>0 else 1/ref.annot.annot_dict[tr]["cds_len"]
                    for tr in count5_pos.columns
                ])
                # normalize by total reads?
                #  * (1e+6/count5_pos.sum(axis=None).sum())
                # idx = tr_norm_term>0
                # count5_pos_norm = count5_pos.values.astype(float)
                count5_pos_norm = count5_pos.values / tr_norm_term[None,:]
            else:
                count5_pos_norm = count5_pos.values
            # count5_pos_norm[ np.isinf(count5_pos_norm)] = 0
            # count5_pos_norm[ np.isnan(count5_pos_norm)] = 0
            count5_pos_norm_fill = np.zeros((np.max(count5_pos.index)-np.min(count5_pos.index)+1, len(count5_pos.columns) ))
            count5_pos_norm_fill[ np.array(count5_pos.index)-np.min(count5_pos.index), : ] = count5_pos_norm
            # for i,tr in enumerate(count5_pos.columns):
            #     count5_pos_norm_fill[ np.array(count5_pos.index)-np.min(count5_pos.index), i ] += count5_pos_norm[:,i]
            coo = coo_matrix(count5_pos_norm_fill,shape=count5_pos_norm_fill.shape)
            dict_col = np.array(list(range( np.min(count5_pos.index), np.max(count5_pos.index)+1 )))
            dict_row = list(count5_pos.columns)
            dict_encode[f'count{end}'] = coo
            dict_encode[f'tr{end}'] = dict_row
            dict_encode[f'pos{end}'] = dict_col
            
    return dict_encode

def calc_norm_ribo_density(
    save_dir,
    load_dir,
    smpls,
    ref,
    read_len_list=[],
    is_length=False,
    load_dir_adjust = '',
    is_norm=True,
    dirname='',
    biotype='',
    full_align=False
):
    if full_align:
        dirname = dirname + '_full'
    save_dir = save_dir / f'norm_ribo_density{dirname}'
    if not save_dir.exists():
        save_dir.mkdir()
    save_dir = Path(save_dir)

    for s in smpls:
        print(f'\ncalculating normalized ribosome density for {s}...')

        obj = mybin.myBinRibo(
            data_dir=ref.data_dir,
            smpl=s,
            sp=ref.sp,
            save_dir=load_dir if load_dir_adjust == '' else load_dir_adjust
        )
        obj.decode()
        df_data,dict_tr = obj.make_df(tr_cnt_thres=-1,is_seq=False)
        if full_align:
            idx_full = df_data['length'] == df_data['read_length']
            df_data = df_data.iloc[ idx_full.values, : ]
            dict_tr = { i[0]:i[1]  for i in list(df_data.groupby('tr_id'))}
        read_len_list_all = np.sort(df_data['read_length'].unique())
        
        for mode in ['start','stop']:
            if not is_length:
                obj = mybin.myBinRiboNorm(
                    smpl=s,
                    sp=ref.sp,
                    save_dir=save_dir,
                    mode=mode,
                    read_len_list=read_len_list,
                    is_length=False,
                    is_norm=True,
                    dirname=''
                )
                if len(read_len_list)>0:
                    print(f'read length {read_len_list[0]} to {read_len_list[1]}...')
                    dict_encode = _calc_norm_ribo_density(
                        df_data,
                        dict_tr,
                        [read_len_list[0],read_len_list[-1]],
                        mode,
                        ref,
                        is_norm)
                else:
                    print(f'all read length...')
                    dict_encode = _calc_norm_ribo_density(
                        df_data,
                        dict_tr,
                        [read_len_list_all[0],read_len_list_all[-1]],
                        mode,
                        ref,
                        is_norm)
                
                obj.encode(dict_encode=dict_encode)

            # for each read length
            else:
                obj = mybin.myBinRiboNorm(
                    smpl=s,
                    sp=ref.sp,
                    save_dir=save_dir,
                    mode=mode,
                    read_len_list=read_len_list,
                    is_length=True,
                    is_norm=True,
                    dirname=''
                )
                df_data_readlen = df_data.groupby('read_length')
                dict_reads = {}
                for read_len in read_len_list_all:
                    print(f'\r{read_len} read length...',end='')
                    dict_reads[read_len] = _calc_norm_ribo_density(
                        df_data_readlen.get_group(read_len),
                        dict_tr,
                        [read_len,read_len],
                        mode,
                        ref,
                        is_norm)
                    if (dict_reads[read_len]['count5'] == []) or (dict_reads[read_len]['count3'] == []):
                        del dict_reads[read_len]
                obj.encode_readlen(dict_reads=dict_reads)
            
def calc_norm_coverage(
    load_dir,
    save_dir,
    threshold_tr_count,
    smpls,
    ref):

    save_dir = save_dir / f'norm_coverage'
    if not save_dir.exists():
        save_dir.mkdir()
    
    for s in smpls:
        print(f'\ncalculating normalized ribosome coverage for {s}...')

        # BAM file
        a = ref.exp_metadata.df_metadata.query(f'sample_name == "{s}"').align_files.iloc[-1]
        infile = pysam.AlignmentFile(ref.data_dir / a)

        # dict_tr
        obj = mybin.myBinRibo(
            data_dir=ref.data_dir,
            smpl=s,
            sp=ref.sp,
            save_dir=load_dir
        )
        obj.decode()
        df_data,dict_tr = obj.make_df(tr_cnt_thres=threshold_tr_count)

        len_tr = len(dict_tr)
        cvgs_cds = {}
        cvgs_3utr = {}
        cvgs_5utr = {}
        for i,(tr,val) in enumerate(dict_tr.items()):
            print(f'\r{i}/{len_tr} transcripts',end='')
            if tr not in ref.annot.annot_dict.keys():
                continue
            # ACGT
            cvge_ = infile.count_coverage(contig=tr)
            cvge = np.array(cvge_[0]) + np.array(cvge_[1]) + np.array(cvge_[2]) + np.array(cvge_[3])
            cds_cnt = infile.count(
                contig=tr,
                start=ref.annot.annot_dict[tr]['start'],
                stop=ref.annot.annot_dict[tr]['stop']+3)
            cvge = cvge / (cds_cnt / ref.annot.annot_dict[tr]['cds_len'])
            # cvge = cvge / (ref.annot.annot_dict[tr]['cds_len'] * total_counts)
            cvge_5utr = cvge[: ref.annot.annot_dict[tr]['start'] ]
            cvge_3utr = cvge[ref.annot.annot_dict[tr]['stop']+3: ]
            cvge_cds = cvge[ ref.annot.annot_dict[tr]['start']:\
                ref.annot.annot_dict[tr]['stop']+3 ]
            cvgs_cds[tr] = cvge_cds
            cvgs_3utr[tr] = cvge_3utr
            cvgs_5utr[tr] = cvge_5utr
        
        my._mysave(save_dir / f'norm_cvgs_cds_{s}.joblib',cvgs_cds)
        my._mysave(save_dir / f'norm_cvgs_3utr_{s}.joblib',cvgs_3utr)
        my._mysave(save_dir / f'norm_cvgs_5utr_{s}.joblib',cvgs_5utr)

def read_stat(
    load_dir:Path,
    ref:my.Ref,
    excel_file:Path,
    readfile_names:list,
    dict_adapters:OrderedDict,
    is_star_antisense = False,
    smpls=[]
):
    
    grayfill = PatternFill(
        start_color='00C0C0C0',
        end_color='00C0C0C0',
        fill_type='solid'
    )
    colnum_int = []
    colnum_perc = []
        
    if (load_dir / 'preprocessing' / 'log_bowtie.txt').exists():
        log_bowtie = load_dir / 'preprocessing' / 'log_bowtie.txt'

    prep_data_file = load_dir / 'prep_data' / 'summary.csv.gz'
    df_prep = pd.read_csv(prep_data_file,header=0,index_col=0)

    # book = openpyxl.load_workbook(excel_file)
    book = openpyxl.Workbook()
    sheet = book['Sheet']
    sheet.title = 'reads_info'

    sheet['A2'] = 'Total input reads'
    colnum_int.append(2)
    sheet['A3'] = 'Adapter trimming by cutadapt'
    
    for i,(adapter,path_adapter) in enumerate(dict_adapters.items()):
        sheet[f'A{i*8+4}'] = f'{adapter} trimming'
        sheet[f'A{i*8+4}'].fill = grayfill
        sheet[f'A{i*8+5}'] = 'Input reads'
        colnum_int.append(i*8+5)
        sheet[f'A{i*8+6}'] = f'Reads with {adapter}'
        colnum_int.append(i*8+6)
        colnum_perc.append(i*8+7)
        sheet[f'A{i*8+8}'] = 'Reads <15 after trimming'
        colnum_int.append(i*8+8)
        colnum_perc.append(i*8+9)
        sheet[f'A{i*8+10}'] = 'Reads used for downstream analysis'
        colnum_int.append(i*8+10)
        colnum_perc.append(i*8+11)

    ii = i*8+12
    sheet[f'A{ii}'] = 'Alignment to rRNA by bowtie'
    sheet[f'A{ii+1}'] = 'Input reads'
    colnum_int.append(ii+1)
    sheet[f'A{ii+2}'] = 'Reads aligned to rRNA'
    colnum_int.append(ii+2)
    colnum_perc.append(ii+3)
    sheet[f'A{ii+4}'] = 'Reads used for downstream analysis'
    colnum_int.append(ii+4)
    colnum_perc.append(ii+5)
    
    if is_star_antisense:
        sheet[f'A{ii+6}'] = 'Alignment to mRNA by STAR'
        sheet[f'A{ii+7}'] = 'Input reads'
        colnum_int.append(ii+7)
        sheet[f'A{ii+8}'] = 'Uniquely mapped reads (sense)'
        colnum_int.append(ii+8)
        colnum_perc.append(ii+9)
        sheet[f'A{ii+10}'] = 'Uniquely mapped & fully aligned reads (sense)'
        colnum_int.append(ii+10)
        colnum_perc.append(ii+11)
        sheet[f'A{ii+12}'] = 'Uniquely mapped reads (antisense)'
        colnum_int.append(ii+12)
        colnum_perc.append(ii+13)
        sheet[f'A{ii+14}'] = 'Uniquely mapped & fully aligned reads (antisense)'
        colnum_int.append(ii+14)
        colnum_perc.append(ii+15)
    
    else:
        sheet[f'A{ii+6}'] = 'Alignment to mRNA by STAR'
        sheet[f'A{ii+7}'] = 'Input reads'
        colnum_int.append(ii+7)
        sheet[f'A{ii+8}'] = 'Uniquely mapped reads'
        colnum_int.append(ii+8)
        colnum_perc.append(ii+9)
        sheet[f'A{ii+10}'] = 'Uniquely mapped & fully aligned reads'
        colnum_int.append(ii+10)
        colnum_perc.append(ii+11)

    alphabet = list(string.ascii_uppercase)

    for j in [3,ii,ii+6]:
        for s in sheet[f'A{j}':f'{alphabet[len(ref.exp_metadata.df_metadata)]}{j}']:
            for ss in s:
                ss.fill = grayfill

    # # initialization
    # for a in alphabet[1:]:
    #     for i in range(1,30):
    #         sheet[f'{a}{i}'] = None

    i = 0
    for ( k,row ),readfile_name in zip(ref.exp_metadata.df_metadata.iterrows(),readfile_names):
        
        if (len(smpls)>0) and (row['sample_name'] not in smpls):
            continue
        
        if is_star_antisense and ('rev' in readfile_name):
            i += 1
            sheet[f'{alphabet[i]}{ii+12}'] = int(df_prep.loc[row['sample_name'],'total_reads'])
            sheet[f'{alphabet[i]}{ii+14}'] = int(df_prep.loc[row['sample_name'],'fully_aligned_reads'])
            continue

        k += 1
        if len(row['barcode'])>1:
            sheet[f'{alphabet[k]}1'] = row['sample_name'] + ' (' + str(i) + '_' + row['barcode'] + ')'
        else:
            sheet[f'{alphabet[k]}1'] = row['sample_name']

        # cutadapt
        for i,(adapter,path_adapter) in enumerate(dict_adapters.items()):
            sheet[f'{alphabet[k]}{i*8+5}'] = f'={alphabet[k]}{i*8+2}'
            sheet[f'{alphabet[k]}{i*8+7}'] = f'={alphabet[k]}{i*8+6}/{alphabet[k]}2'
            sheet[f'{alphabet[k]}{i*8+9}'] = f'={alphabet[k]}{i*8+8}/{alphabet[k]}2'
            sheet[f'{alphabet[k]}{i*8+11}'] = f'={alphabet[k]}{i*8+10}/{alphabet[k]}2'

            with open(path_adapter % readfile_name, 'r') as f:
                for line in f:
                    if line.startswith('Total reads processed:'):
                        if i==0:
                            sheet[f'{alphabet[k]}2'] = int(line.split(':')[-1].strip().replace(',',''))
                    if line.startswith('Reads with adapters:'):
                        sheet[f'{alphabet[k]}{i*8+6}'] = int(line.split(':')[-1].split('(')[0].strip().replace(',',''))
                    if line.startswith('Reads that were too short:'):
                        sheet[f'{alphabet[k]}{i*8+8}'] = int(line.split(':')[-1].split('(')[0].strip().replace(',',''))
                    if line.startswith('Reads written (passing filters):'):
                        sheet[f'{alphabet[k]}{i*8+10}'] = int(line.split(':')[-1].split('(')[0].strip().replace(',',''))
        
        # bowtie
        ii = i*8+12
        sheet[f'{alphabet[k]}{ii+1}'] = f'={alphabet[k]}{ii-2}'
        sheet[f'{alphabet[k]}{ii+3}'] = f'={alphabet[k]}{ii+2}/{alphabet[k]}{ii+1}'
        sheet[f'{alphabet[k]}{ii+5}'] = f'={alphabet[k]}{ii+4}/{alphabet[k]}{ii+1}'
        
        # STAR
        sheet[f'{alphabet[k]}{ii+9}'] = f'={alphabet[k]}{ii+8}/{alphabet[k]}{ii+7}'
        sheet[f'{alphabet[k]}{ii+11}'] = f'={alphabet[k]}{ii+10}/{alphabet[k]}{ii+7}'
        if is_star_antisense:
            sheet[f'{alphabet[k]}{ii+13}'] = f'={alphabet[k]}{ii+12}/{alphabet[k]}{ii+7}'
            sheet[f'{alphabet[k]}{ii+15}'] = f'={alphabet[k]}{ii+14}/{alphabet[k]}{ii+7}'
        
        
        if (load_dir / 'preprocessing' / 'log_bowtie.txt').exists():
            with open(log_bowtie,'r') as f:
                flag = False
                for line in f:
                    if line.startswith(readfile_name):
                        flag = True
                    if flag:
                        if line.startswith('# reads processed:'):
                            assert int(line.split(':')[-1].strip()) == sheet[f'{alphabet[k]}{ii-2}'].value
                        if line.startswith('# reads with at least one reported alignment:'):
                            sheet[f'{alphabet[k]}{ii+2}'] = int(line.split(':')[-1].split('(')[0].strip())
                            flag = False
            sheet[f'{alphabet[k]}{ii+4}'] = f'={alphabet[k]}{ii+1}-{alphabet[k]}{ii+2}'
            
        else:
            log_star = load_dir / 'preprocessing' / f'STAR_align_{readfile_name}_Log.final.out'
            if not log_star.exists():
                raise Exception(f'{log_star} does not exist')
            with open(log_star,'r') as f:
                for line in f:
                    if line.strip().startswith('Number of input reads'):
                        sheet[f'{alphabet[k]}{ii+4}'] = int(line.strip().split('|\t')[1])
            sheet[f'{alphabet[k]}{ii+2}'] = f'={alphabet[k]}{ii+1}-{alphabet[k]}{ii+4}'
        
        sheet[f'{alphabet[k]}{ii+7}'] = f'={alphabet[k]}{ii+4}'
        sheet[f'{alphabet[k]}{ii+8}'] = int(df_prep.loc[row['sample_name'],'total_reads'])
        sheet[f'{alphabet[k]}{ii+10}'] = int(df_prep.loc[row['sample_name'],'fully_aligned_reads'])

        # format
        for j in colnum_int:
            # fmt_now = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            fmt_now = u'#,##0'
            sheet[f'{alphabet[k]}{j}'].number_format = fmt_now
        for j in colnum_perc:
            sheet[f'{alphabet[k]}{j}'].number_format = numbers.FORMAT_PERCENTAGE_00

    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length
    book.save(excel_file)

def combine_datasets(
    save_dir,
    fname,
    ref,
    ribobin_path:list
):
    save_dir = save_dir / 'combine_datasets'
    if not save_dir.exists():
        save_dir.mkdir()
    
    df = []
    for p in ribobin_path:
        s = re.search(r'dat_(.*).bin',Path(p).name).group(1)
        obj = mybin.myBinRibo(
            data_dir=ref.data_dir,
            smpl=s,
            sp=ref.sp,
            save_dir=Path(p).parent
        )
        obj.decode()
        df_data,dict_tr = obj.make_df(tr_cnt_thres=-1,is_cds=True,is_frame=True,is_seq=True)
        if len(df)>0:
            df = pd.concat([df, df_data], axis=0)
        else:
            df = df_data
    
    df_ = df.sort_values(['tr_id','cut5'])

    obj = mybin.myBinRibo(
        data_dir=ref.data_dir,
        smpl=fname,
        sp=ref.sp,
        save_dir=save_dir
    )
    obj.encode_df(df=df_)
    print("hoge")


def combine_datasets_gene(
    save_dir,
    fname,
    ref,
    ribobin_path:list
):
    save_dir = save_dir / 'combine_datasets'
    if not save_dir.exists():
        save_dir.mkdir()
    
    if '_gene' in ribobin_path[0].stem:
        mode = 'gene'
    else:
        mode = 'tr'
    
    df = []
    for p in ribobin_path:
        s = re.search(r'dat_(.*).bin',Path(p).stem).group(1)
        obj = mybin.myBinRiboGenome(
            data_dir=ref.data_dir,
            smpl=s,
            sp=ref.sp,
            save_dir=Path(p).parent
        )
        obj.decode(data_path=p,mode=mode)
        df_data,_ = obj.make_df(mode=mode,is_seq=True)
        if len(df)>0:
            df = pd.concat([df, df_data], axis=0)
        else:
            df = df_data
    
    df_ = df.sort_values(['tr_id','cut5'])

    obj = mybin.myBinRiboGenome(
        data_dir=ref.data_dir,
        smpl=fname,
        sp=ref.sp,
        save_dir=save_dir
    )
    obj.encode_df(mode=mode,df=df_)
    print("hoge")


def map_tr_to_genome(
    path_bam:Path,
    ref:my.RefGenome,
    biotype:str,
    save_dir:Path
    ):
        save_dir = save_dir / 'preprocessing'
        if not save_dir.exists():
            save_dir.mkdir()

        assert biotype in biotype_list

        header = {'HD':{},'SQ':[]}
        dict_chr_dict = {};i=0
        with open(ref.ref_dir_genome / 'bam_header_genome', 'r') as f:
            for line in f:
                l = line.strip().split('\t')
                if l[0] == '@HD':
                    lll = l[1].split(':')
                    header[l[0][1:]][lll[0]] = float(lll[1])
                    lll = l[2].split(':')
                    header[l[0][1:]][lll[0]] = lll[1]
                elif l[0] == '@SQ':
                    tmp = {}
                    for ll in l[1:]:
                        lll = ll.split(':')
                        if lll[0] == 'LN':
                            tmp[lll[0]] = int(lll[1])
                        else:
                            tmp[lll[0]] = lll[1]
                    header[l[0][1:]].append(tmp)
                if l[0] == '@SQ':
                    dict_chr_dict[l[1].split(':')[1]] = i
                    i += 1
        
        infile = pysam.AlignmentFile(path_bam)
        
        annot_genome = ref.get_annot_gene(biotype)
        # seq_gene = ref.get_seq_gene()

        outfile = pysam.AlignmentFile(save_dir / f'{path_bam.stem}_togenome.bam', 'wb', header=header)

        n_genes = len(annot_genome.annot_dict)
        for i,(gene,dtg) in enumerate(annot_genome.annot_dict.items()):
            print(f'\r{i}/{n_genes} genes...',end='')

            tr = dtg['tr_id']
            if tr not in infile.references:
                continue
            if infile.count(tr) == 0:
                continue
            dtg = annot_genome.annot_dict[ gene ]
            # len_gene = np.sum([x[1]-x[0]+1 for x in dtg['regions']])
            # assert len(seq_gene[dt['gene_id']]) == len_gene

            reads_tr_iter = [r for r in infile.fetch(tr)]

            if dtg['strand'] == '+':
                exon_starts = np.array([x[0] for x in dtg['regions'] if x[2]==1])
                exon_lengths = np.cumsum([0]+[x[1]-x[0]+1 for x in dtg['regions'] if x[2]==1])
            elif dtg['strand'] == '-':
                exon_starts = np.array([x[0] for x in dtg['regions'][::-1] if x[2]==1])
                exon_lengths = np.cumsum([0]+[x[1]-x[0]+1 for x in dtg['regions'][::-1] if x[2]==1])
                
            for r in reads_tr_iter:
                
                whereitis = [
                    np.where((r.reference_start - exon_lengths) >= 0)[0][-1],
                    np.where((r.reference_end - exon_lengths) >= 0)[0][-1]
                ]

                r.reference_id = dict_chr_dict[str(dtg['chr'])]
                if dtg['strand'] == '+':
                    r.reference_start = (r.reference_start - exon_lengths[whereitis[0]]) + exon_starts[whereitis[0]] + dtg['start'] -1
                elif dtg['strand'] == '-':
                    if whereitis[1] == len(exon_starts):
                        r.reference_start = (exon_lengths[-1] - r.reference_end) + exon_starts[-1] + dtg['start'] -1
                    else:
                        r.reference_start = (exon_lengths[whereitis[1]+1] - r.reference_end) + exon_starts[whereitis[1]] + dtg['start'] -1
                    r.is_reverse = True
                    r.query_sequence = my._rev(r.query_sequence)
                    r.cigartuples = r.cigartuples[::-1]
                
                # # check
                # if dtg['strand'] == '+':
                #     print(seq_gene[gene][ r.reference_start - dtg['start'] : r.reference_end - dtg['start'] ])
                #     print(r.query_alignment_sequence)
                #     print(r.reference_start - dtg['start'])
                #     print(seq_gene[gene].find(r.query_alignment_sequence))
                    
                # elif dtg['strand'] == '-':
                #     print(my._rev(seq_gene[gene][ r.reference_start - dtg['start'] : r.reference_end - dtg['start'] ]))
                #     print(r.query_alignment_sequence)
                #     print(r.reference_start - dtg['start'])
                #     print(seq_gene[gene].find(my._rev(r.query_alignment_sequence)))

                outfile.write(r)

        outfile.close()

        pysam.sort(
            "-o", 
            (save_dir / f'{path_bam.stem}_togenome_sorted.bam').as_posix(),
            (save_dir / f'{path_bam.stem}_togenome.bam').as_posix()
        )
        pysam.index(
            (save_dir / f'{path_bam.stem}_togenome_sorted.bam').as_posix()
        )
